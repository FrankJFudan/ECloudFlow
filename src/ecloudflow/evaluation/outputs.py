"""Typed tabular, workbook, JSON, and rank-ordered SDF writers."""

from __future__ import annotations

import csv
import json
import math
from collections.abc import Sequence
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from rdkit import Chem

from ecloudflow.evaluation.ranking import RankedMolecule
from ecloudflow.sampling.results import GenerationRecord


@dataclass(frozen=True)
class OutputBundle:
    """List every artifact emitted by :func:`write_ranked_outputs`.

    :param paths: Stable tuple of generated artifact paths.
    :param rows: Number of formally ranked rows.
    :param failed_rows: Number of unranked/failed rows.
    :return: Immutable output manifest.
    :rtype: OutputBundle
    """

    paths: tuple[Path, ...]
    rows: int = 0
    failed_rows: int = 0

    @property
    def files(self) -> tuple[Path, ...]:
        """Return an alias for :attr:`paths` used by CLI/report callers."""
        return self.paths


def write_ranked_outputs(
    ranked: Sequence[RankedMolecule],
    failed: Sequence[GenerationRecord],
    output_dir: str | Path,
) -> OutputBundle:
    """Write complete ranked and failed molecule artifacts.

    :param ranked: Rows in final docking rank order.
    :param failed: Generation records without a usable docking score.
    :param output_dir: Destination directory; created when absent.
    :return: Paths and row counts for CSV, Parquet, Excel, JSON, and SDF files.
    :rtype: OutputBundle
    :raises ValueError: If rows contain malformed ranking identities.

    The writer never recomputes docking or molecular properties.  It serializes
    the values already present in the typed rows and retains failed records in
    a separate table.  SDF records follow exactly the same order as the ranked
    table and carry the formal molecule ID as an SD property.
    """
    destination = Path(output_dir)
    destination.mkdir(parents=True, exist_ok=True)
    ranked_rows = [row.as_dict() for row in ranked]
    failed_rows = [_failed_row(record) for record in failed]
    aggregate_rows = _aggregate_rows(ranked_rows)

    csv_path = destination / "samples.csv"
    failed_csv_path = destination / "failed.csv"
    parquet_path = destination / "samples.parquet"
    excel_path = destination / "summary.xlsx"
    sdf_path = destination / "ranked.sdf"
    summary_path = destination / "summary.json"

    _write_csv(csv_path, ranked_rows)
    _write_csv(failed_csv_path, failed_rows)
    _write_parquet(parquet_path, ranked_rows)
    _write_excel(excel_path, ranked_rows, failed_rows, aggregate_rows)
    _write_ranked_sdf(sdf_path, ranked)
    _atomic_json(
        summary_path,
        {
            "ranked_count": len(ranked_rows),
            "failed_count": len(failed_rows),
            "ranked": ranked_rows,
            "failed": failed_rows,
            "aggregate": aggregate_rows,
        },
    )
    return OutputBundle(
        paths=(
            csv_path,
            parquet_path,
            excel_path,
            sdf_path,
            summary_path,
            failed_csv_path,
        ),
        rows=len(ranked_rows),
        failed_rows=len(failed_rows),
    )


def _failed_row(record: GenerationRecord) -> dict[str, Any]:
    """Convert an unranked generation record to the failure table schema."""
    properties = dict(record.properties)
    return {
        "rank": None,
        "molecule_id": None,
        "pocket_id": properties.get("pocket_id"),
        "temporary_id": record.temporary_id,
        "canonical_smiles": record.canonical_smiles,
        "isomeric_smiles": record.canonical_smiles,
        "sa": _first(properties, "sa", "sa_score", "synthetic_accessibility"),
        "qed": _first(properties, "qed", "QED"),
        "docking_score": None,
        "generation_status": properties.get("status", "dock_failed"),
        "raw_path": str(record.raw_path) if record.raw_path else None,
        "relaxed_path": str(record.relaxed_path) if record.relaxed_path else None,
        "seed": record.seed,
        "model_checkpoint_hash": record.model_checkpoint_hash,
    }


def _aggregate_rows(rows: Sequence[dict[str, Any]]) -> list[dict[str, Any]]:
    """Compute standard descriptive statistics without returning NaN JSON."""
    output: list[dict[str, Any]] = []
    for metric, aliases in (
        ("sa", ("sa", "sa_score")),
        ("qed", ("qed",)),
        ("docking_score", ("docking_score", "vina_score")),
    ):
        values = []
        for row in rows:
            value = _first(row, *aliases)
            if value is not None:
                value = float(value)
                if not math.isfinite(value):
                    raise ValueError(f"aggregate metric {metric!r} must be finite")
                values.append(value)
        stats = _summary(values)
        for name, value in stats.items():
            output.append({"metric": metric, "statistic": name, "value": value})
    return output


def _summary(values: Sequence[float]) -> dict[str, float | int | None]:
    """Return count/mean/std/quantile statistics for one metric."""
    if not values:
        return {
            "count": 0,
            "mean": None,
            "std": None,
            "median": None,
            "min": None,
            "max": None,
            "q1": None,
            "q3": None,
        }
    ordered = sorted(values)
    mean = sum(ordered) / len(ordered)
    variance = (
        sum((value - mean) ** 2 for value in ordered) / (len(ordered) - 1)
        if len(ordered) > 1
        else 0.0
    )
    return {
        "count": len(ordered),
        "mean": mean,
        "std": math.sqrt(variance),
        "median": _quantile(ordered, 0.5),
        "min": ordered[0],
        "max": ordered[-1],
        "q1": _quantile(ordered, 0.25),
        "q3": _quantile(ordered, 0.75),
    }


def _quantile(values: Sequence[float], probability: float) -> float:
    """Compute a deterministic linear-interpolated quantile."""
    if len(values) == 1:
        return float(values[0])
    position = probability * (len(values) - 1)
    lower = int(position)
    upper = min(lower + 1, len(values) - 1)
    fraction = position - lower
    return float(values[lower] + fraction * (values[upper] - values[lower]))


def _write_csv(path: Path, rows: Sequence[dict[str, Any]]) -> None:
    """Write a rectangular CSV with a stable union of column names."""
    keys = _keys(rows)
    temporary = path.with_name(path.name + ".partial")
    with temporary.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=keys, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    temporary.replace(path)


def _write_parquet(path: Path, rows: Sequence[dict[str, Any]]) -> None:
    """Write typed Parquet when pandas/Arrow are available."""
    try:
        import pandas as pd

        temporary = path.with_name(path.name + ".partial")
        pd.DataFrame(rows).to_parquet(temporary, index=False)
        temporary.replace(path)
    except ImportError as error:
        raise RuntimeError("Parquet output requires pandas and pyarrow.") from error


def _write_excel(
    path: Path,
    ranked_rows: Sequence[dict[str, Any]],
    failed_rows: Sequence[dict[str, Any]],
    aggregate_rows: Sequence[dict[str, Any]],
) -> None:
    """Write the required ranked/failed/aggregate workbook sheets."""
    temporary = path.with_name(path.name + ".partial")
    try:
        import pandas as pd

        with pd.ExcelWriter(temporary) as writer:
            pd.DataFrame(ranked_rows).to_excel(writer, sheet_name="ranked", index=False)
            pd.DataFrame(failed_rows).to_excel(writer, sheet_name="failed", index=False)
            pd.DataFrame(aggregate_rows).to_excel(
                writer, sheet_name="aggregate", index=False
            )
    except ImportError:
        try:
            from openpyxl import Workbook
        except ImportError as error:
            raise RuntimeError("Excel output requires pandas or openpyxl.") from error
        workbook = Workbook()
        _append_sheet(workbook.active, ranked_rows, "ranked")
        _append_sheet(workbook.create_sheet(), failed_rows, "failed")
        _append_sheet(workbook.create_sheet(), aggregate_rows, "aggregate")
        workbook.save(temporary)
    temporary.replace(path)


def _append_sheet(sheet: Any, rows: Sequence[dict[str, Any]], title: str) -> None:
    """Populate one openpyxl sheet from dictionaries."""
    sheet.title = title
    keys = _keys(rows)
    sheet.append(keys)
    for row in rows:
        sheet.append([row.get(key) for key in keys])


def _write_ranked_sdf(path: Path, ranked: Sequence[RankedMolecule]) -> None:
    """Write molecules in rank order with required SD properties."""
    temporary = path.with_name(path.name + ".partial")
    writer = Chem.SDWriter(str(temporary))
    try:
        for row in ranked:
            molecule = row.molecule
            if molecule is None:
                molecule = Chem.MolFromSmiles(row.canonical_smiles)
            if molecule is None:
                raise ValueError(
                    f"cannot reconstruct SDF molecule {row.canonical_smiles}"
                )
            molecule = Chem.Mol(molecule)
            molecule.SetProp("molecule_id", row.molecule_id)
            molecule.SetProp("rank", str(row.rank))
            molecule.SetProp("canonical_isomeric_smiles", row.canonical_smiles)
            molecule.SetProp("docking_score", str(row.docking_score))
            if row.qed is not None:
                molecule.SetProp("QED", str(row.qed))
            if row.sa_score is not None:
                molecule.SetProp("SA", str(row.sa_score))
            writer.write(molecule)
    finally:
        writer.close()
    temporary.replace(path)


def _atomic_json(path: Path, payload: Any) -> None:
    """Write JSON after recursively replacing non-finite values."""
    temporary = path.with_name(path.name + ".partial")
    temporary.write_text(
        json.dumps(_json_safe(payload), indent=2, sort_keys=True), encoding="utf-8"
    )
    temporary.replace(path)


def _keys(rows: Sequence[dict[str, Any]]) -> list[str]:
    """Return deterministic union keys while preserving schema order."""
    keys: list[str] = []
    for row in rows:
        for key in row:
            if key not in keys:
                keys.append(key)
    return keys or ["status"]


def _first(values: dict[str, Any], *names: str) -> Any:
    """Return the first present non-null alias."""
    for name in names:
        if values.get(name) is not None:
            return values[name]
    return None


def _json_safe(value: Any) -> Any:
    """Convert nested values to JSON-safe primitives."""
    if value is None or isinstance(value, (str, int, bool)):
        return value
    if isinstance(value, float):
        return value if math.isfinite(value) else None
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, dict):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, (tuple, list)):
        return [_json_safe(item) for item in value]
    return str(value)


__all__ = ["OutputBundle", "write_ranked_outputs"]
